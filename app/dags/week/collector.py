import os
import re
import sys
import json
import pandas
import pickle
import requests

from io import BytesIO
from typing import Tuple, List, Dict, Any, Optional, Union, Callable
from colour import Color
from pathlib import Path
from httplib2 import Http
from openpyxl import load_workbook
from datetime import time, date, datetime
from apiclient import discovery
from xlsxwriter import Workbook
from urllib.parse import urlparse, parse_qsl
from transliterate import slugify
from googleapiclient.discovery import Resource as GoogleAPIClientResource
from openpyxl.worksheet.worksheet import Worksheet
from oauth2client.service_account import ServiceAccountCredentials

from airflow import DAG
from airflow.models import Variable
from airflow.operators.python import PythonOperator

try:
    sys.path.append(Variable.get("APP_FOLDER", None))
except KeyError:
    pass

from config import DATA_FOLDER, CREDENTIALS_FILE
from app.analytics import pickle_loader
from app.dags.decorators import log_execution_time


DATA_PATH = Path(DATA_FOLDER) / "week"
os.makedirs(DATA_PATH, exist_ok=True)


def parse_slug(value: str) -> str:
    if str(value) == "" or pandas.isna(value):
        return pandas.NA
    return slugify(str(value), "ru").replace("-", "_")


def parse_lead_id(value: str) -> int:
    if pandas.isna(value):
        return pandas.NA
    lead_id = pandas.NA
    if str(value).startswith("https://neuraluniversity.amocrm.ru/leads/detail/"):
        match = re.search(r"^(\d+)", value[48:])
        lead_id = int(match.group(1))
    return lead_id


def parse_str(value: str) -> str:
    if pandas.isna(value):
        return pandas.NA
    value = re.sub(r"\s+", " ", str(value).strip())
    if value == "" or value == "-":
        value = pandas.NA
    return value


def parse_str_with_empty(value: str) -> str:
    if pandas.isna(value):
        return pandas.NA
    value = re.sub(r"\s+", " ", str(value).strip())
    return value


def parse_int(value: str) -> int:
    value = parse_float(value)
    if pandas.isna(value):
        return pandas.NA
    return round(value)


def parse_float(value: str) -> float:
    if pandas.isna(value):
        return pandas.NA
    value = re.sub(r"\s", "", str(value))
    if not re.search(r"^-?\d+\.?\d*$", str(value)):
        return pandas.NA
    return float(value)


def parse_date(value: str) -> date:
    if pandas.isna(value):
        return pandas.NA
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    match = re.search(r"^(\d{1,2})\.(\d{1,2})\.(\d{4}).*$", str(value))
    if not match:
        return pandas.NA
    groups = list(match.groups())
    if len(groups[0]) == 1:
        groups[0] = f"0{groups[0]}"
    if len(groups[1]) == 1:
        groups[1] = f"0{groups[1]}"
    return date.fromisoformat("-".join(list(reversed(groups))))


def parse_url(value: str) -> str:
    value = parse_str(value)
    if pandas.isna(value):
        return pandas.NA
    url = urlparse(value)
    if url.scheme and url.netloc and url.path:
        return value
    return pandas.NA


def parse_lead_url(value: str) -> str:
    value = parse_url(value)
    if pandas.isna(value):
        return pandas.NA
    url = urlparse(value)._replace(params=None, query=None, fragment=None).geturl()
    if isinstance(url, bytes):
        url = url.decode("utf-8")
    return url


def parse_da_net(value: str) -> bool:
    value = str(value).strip().lower()
    if value == "да":
        return True
    if value == "нет":
        return False
    return pandas.NA


def parse_cell_data(cell: Dict[str, Dict[str, Any]]) -> str:
    value = cell.get("formattedValue", "") or ""
    if value.strip().lower() in ["о", "o"]:
        value = ""
    color = cell.get("effectiveFormat", {}).get("backgroundColor")
    lead_id = parse_lead_id(value)
    if not pandas.isna(lead_id) and (
        not color or Color(rgb=list(color.values())[:3]).get_hex_l() != "#98e098"
    ):
        value = ""
    return value


def parse_row_data(row: List[Dict[str, dict]]) -> List[str]:
    return list(map(parse_cell_data, row.get("values")))


def count_query_params_concurrency(target: dict, value: dict) -> int:
    columns = list(set(list(target.keys()) + list(value.keys())))
    return len(
        list(
            filter(
                lambda item: target.get(item) == value.get(item)
                and target.get(item) is not None
                and value.get(item) is not None,
                columns,
            )
        )
    )


def compare_exact_params(params: List[str], target: dict, value: dict) -> bool:
    output = False
    columns = list(set(list(target.keys()) + list(value.keys())))
    for param in params:
        if param in columns and target.get(param) == value.get(param):
            output = True
            break
    return output


def matched_url(
    target_query: Dict[str, str], matched_query_dict: Dict[str, str]
) -> List[int]:
    target_query = dict(
        map(
            lambda item: (
                item[0],
                re.sub(
                    r"\s+",
                    " ",
                    re.sub(r"\++", " ", re.sub(r"(#[^#]+)$", "", item[1])),
                ),
            ),
            target_query.items(),
        )
    )
    matched_query_dict = dict(
        map(
            lambda row: (
                row[0],
                dict(
                    map(
                        lambda item: (
                            item[0],
                            re.sub(
                                r"\s+",
                                " ",
                                re.sub(r"\++", " ", re.sub(r"(#[^#]+)$", "", item[1])),
                            ),
                        ),
                        row[1].items(),
                    )
                ),
            ),
            matched_query_dict.items(),
        )
    )
    exact_match = dict(
        filter(lambda item: item[1] == target_query, matched_query_dict.items())
    )
    matched_keys = list(exact_match.keys())
    if len(matched_keys):
        return matched_keys

    exact_match_params = dict(
        filter(
            lambda item: compare_exact_params(
                ["yclid", "rs", "roistat"], target_query, item[1]
            ),
            matched_query_dict.items(),
        )
    )
    exact_match_params_keys = list(exact_match_params.keys())
    if len(exact_match_params_keys):
        return exact_match_params_keys

    count_concurrency = dict(
        map(
            lambda item: (
                item[0],
                count_query_params_concurrency(target_query, item[1]),
            ),
            matched_query_dict.items(),
        )
    )
    max_concurrency = max(count_concurrency.values())
    if max_concurrency > 0:
        return list(
            dict(
                filter(
                    lambda item: item[1] == max_concurrency,
                    count_concurrency.items(),
                )
            ).keys()
        )
    else:
        return []


def detect_lead(item: pandas.Series, rows: pandas.DataFrame) -> Optional[pandas.Series]:
    url = item["amo"]
    target = item["target_link"]
    if pandas.isna(target):
        return

    target_parsed = urlparse(target, allow_fragments=False)
    target_query = dict(parse_qsl(target_parsed.query))
    target_short = target_parsed._replace(
        params=None, query=None, fragment=None
    ).geturl()

    if not target_parsed.netloc:
        return

    lead: pandas.Series = None

    amo = rows[(rows["amo_current"] == url) | (rows["amo_main"] == url)].sort_values(
        by=["date"]
    )
    matched: pandas.DataFrame = amo[amo["target_short"] == target_short]
    if len(matched):
        matched_query = dict(
            map(
                lambda item: (
                    item[0],
                    dict(
                        parse_qsl(
                            urlparse(
                                item[1]["traffic_channel"], allow_fragments=False
                            ).query
                        )
                    ),
                ),
                matched.iterrows(),
            )
        )
        indexes = matched_url(target_query, matched_query)
        if len(indexes):
            matched_leads = matched.loc[indexes]
            matched_leads["updated_at"] = matched_leads["updated_at"].apply(
                lambda item: datetime.fromtimestamp(item)
                if isinstance(item, int)
                else item
            )
            matched_leads.sort_values(by="updated_at", ascending=False, inplace=True)
            lead = matched_leads.iloc[0]

    if lead is None:
        amo = rows[
            (
                (rows["amo_current"] == "")
                | rows["amo_current"].isna()
                | rows["amo_current"].isnull()
            )
            & (
                (rows["amo_main"] == "")
                | rows["amo_main"].isna()
                | rows["amo_main"].isnull()
            )
        ]
        matched = amo[amo["target_short"] == target_short]
        if len(matched):
            matched_query = dict(
                map(
                    lambda item: (
                        item[0],
                        dict(
                            parse_qsl(
                                urlparse(
                                    item[1]["traffic_channel"], allow_fragments=False
                                ).query
                            )
                        ),
                    ),
                    matched.iterrows(),
                )
            )
            indexes = matched_url(target_query, matched_query)
            if len(indexes):
                matched_leads = matched.loc[indexes]
                matched_leads["updated_at"] = matched_leads["updated_at"].apply(
                    lambda item: datetime.fromtimestamp(item)
                    if isinstance(item, int)
                    else item
                )
                matched_leads.sort_values(
                    by="updated_at", ascending=False, inplace=True
                )
                lead = matched_leads.iloc[0]

    return lead


def slugify_columns(columns: List[str]) -> List[str]:
    return list(map(parse_slug, columns))


def rename_so_columns(value: str) -> str:
    map_values = {
        "menedzher": "Менеджер",
        "gruppa": "Группа",
        "sdelka": "Сделка",
        "fio_klienta": "ФИО клиента",
        "email": "Email",
        "telefon": "Телефон",
        "data_zoom": "Дата Zoom",
        "data_so": "Дата SO",
        "do_ili_posle_zoom": "До или после Zoom",
        "id_oplaty_": "ID оплаты ",
        "data_oplaty_": "Дата оплаты ",
        "summa_oplaty_": "Сумма оплаты ",
    }
    if value in map_values.keys():
        return map_values.get(value)

    match = re.search(r"^(\D+)(\d+)$", value)
    if match and match.group(1) in map_values.keys():
        return f"{map_values.get(match.group(1))}{match.group(2)}"

    return value


def get_spreadsheet_url(name: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{name}"


def read_zoom_day(
    worksheet: Worksheet, current_date: date
) -> Tuple[Optional[pandas.DataFrame], Optional[pandas.DataFrame]]:
    print(f'    - Reading day {current_date.strftime("%d.%m.%Y")}')
    values = []
    for row in worksheet.rows:
        row = list(map(lambda cell: cell.value, row))
        if not len(list(filter(None, row))):
            continue
        values.append(row)

    if not len(values):
        return

    columns = values[0]
    columns[0] = "manager"
    values_data = pandas.DataFrame(values[1:], columns=columns)
    prepare_data = []
    managers = []
    group = ""
    for index, row in values_data.iterrows():
        manager = parse_str(row["manager"])
        if pandas.isna(manager):
            continue
        group_match = re.match(r"Группа\s(\S+)", manager)
        if group_match:
            group = group_match.group(1)
            continue
        managers.append({"manager": manager, "group": group, "date": current_date})
        for column, value in row.items():
            if not isinstance(column, time):
                continue
            value = parse_lead_id(value)
            if pandas.isna(value):
                continue
            prepare_data.append(
                {
                    "manager": manager,
                    "date": current_date,
                    "time": column,
                    "lead": value,
                }
            )

    return pandas.DataFrame(managers), pandas.DataFrame(prepare_data)


def read_zoom_month(
    spreadsheet_id: str, month: date
) -> Tuple[pandas.DataFrame, Optional[pandas.DataFrame]]:
    print(
        f'  - Reading month {month.strftime("%m.%Y")} from url {get_spreadsheet_url(spreadsheet_id)}'
    )
    values = None
    managers = pandas.DataFrame(columns=["manager", "group"])
    response = requests.get(
        f"https://spreadsheets.google.com/feeds/download/spreadsheets/Export?key={spreadsheet_id}&exportFormat=xlsx"
    )
    try:
        xlsx = load_workbook(filename=BytesIO(response.content))
    except Exception:
        return managers, values
    for worksheet in xlsx.worksheets:
        if worksheet.sheet_state.lower() == "hidden":
            continue
        match_day = re.match(r"^[а-яА-Я]{2},?\s*(\d+)$", worksheet.title.strip())
        if not match_day:
            continue
        try:
            current_date = month.replace(day=int(match_day.group(1)))
        except ValueError:
            continue
        managers_day, values_day = read_zoom_day(worksheet, current_date)
        if managers_day is not None:
            managers = pandas.concat([managers, managers_day]).reset_index(drop=True)
        if values_day is None:
            continue
        if values is None:
            values = pandas.DataFrame()
        values = pandas.concat([values, values_day]).reset_index(drop=True)
    return managers, values


def read_payments(
    service: GoogleAPIClientResource, spreadsheet_id: str
) -> pandas.DataFrame:
    spreadsheets = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    payments_columns = ["manager_id", "lead", "date", "profit"]
    output = pandas.DataFrame(columns=payments_columns)
    for sheet in spreadsheets.get("sheets"):
        title = sheet.get("properties").get("title")
        if title == "Все оплаты":
            values = (
                service.spreadsheets()
                .values()
                .get(spreadsheetId=spreadsheet_id, range=title, majorDimension="ROWS")
                .execute()
            ).get("values")
            columns = slugify_columns(values[0])
            values = values[1:]
            payments = pandas.DataFrame(
                list(map(lambda item: dict(zip(columns, item)), values))
            )
            payments["lead"] = payments["ssylka_na_amocrm"].apply(parse_lead_id)
            payments["manager_id"] = payments["menedzher"].apply(parse_slug)
            payments["date"] = payments["data_zoom"].apply(parse_date)
            payments["profit"] = payments["summa_vyruchki"].apply(parse_int).fillna(0)
            payments = payments[payments_columns]
            payments = payments[
                (
                    ~(
                        payments["date"].isna()
                        | payments["lead"].isna()
                        | payments["manager_id"].isna()
                    )
                )
                & (payments["profit"] > 0)
            ]
            output = pandas.DataFrame(
                list(
                    map(
                        lambda item: dict(
                            zip(
                                payments_columns,
                                list(item[0]) + [item[1]["profit"].sum()],
                            )
                        ),
                        payments.groupby(by=["manager_id", "lead", "date"]),
                    )
                )
            )
    return output


def get_google_service() -> GoogleAPIClientResource:
    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        CREDENTIALS_FILE,
        [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    http_auth = credentials.authorize(Http())
    return discovery.build("sheets", "v4", http=http_auth)


def get_spreadsheets_month(service: GoogleAPIClientResource) -> pandas.DataFrame:
    def parse_row(row: pandas.Series):
        url = urlparse(row["ssylki"])
        if not url.path.startswith("/spreadsheets/d/"):
            raise ValueError(
                f'Incorrect url {row["ssylki"]} for month {row["mesjatsgod"]}'
            )
        match_id = re.match(r"^([^\/]+).*$", url.path[16:])
        if not match_id:
            raise ValueError(
                f'Incorrect url {row["ssylki"]} for month {row["mesjatsgod"]}'
            )
        row["hash"] = match_id.group(1)
        row["month"] = "".join(list(reversed(row["mesjatsgod"].split("."))))
        return row

    spreadsheet_id = "1H30vy-_-Zeqw_IdGFUMLshsxKw9TwCfqGh6fDXH8GmU"
    print(f"- Reading spreadsheets from url {get_spreadsheet_url(spreadsheet_id)}")
    values = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range="Основные таблицы",
            majorDimension="ROWS",
        )
        .execute()
    ).get("values")

    data = pandas.DataFrame(values[1:], columns=slugify_columns(values[0]))
    data = data.apply(parse_row, axis=1)
    return data[["hash", "month"]]


PAYMENTS_COLUMNS = {
    "ФИО Плательщика": parse_str,
    "ФИО студента": parse_str,
    "Почта": parse_str,
    "Телефон": parse_str,
    "Ссылка на amocrm": parse_lead_url,
    "Менеджер": parse_str,
    "ГР": parse_int,
    "Сумма оплаченная клиентом": parse_float,
    "Сумма выручки": parse_float,
    "Номер заказа": parse_str,
    "Источник оплаты": parse_str,
    "Дата создания сделки": parse_date,
    "Дата последней заявки (платной)": parse_date,
    "Дата оплаты": parse_date,
    "Воронка": parse_str,
    "Месяц / Доплата": parse_str,
    "Дата ZOOM": parse_date,
    "Дата звонка": parse_date,
    "Курс": parse_str,
    "УИИ / УИИ Терра / Терра ЭйАй": parse_str,
    "Стажировка": parse_da_net,
    "Целевая ссылка": parse_url,
}


@log_execution_time("get_payments")
def get_payments():
    service = get_google_service()
    values = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId="1C4TnjTkSIsHs2svSgyFduBpRByA7M_i2sa6hrsX84EE",
            range="Все оплаты",
            majorDimension="ROWS",
        )
        .execute()
    ).get("values")
    columns_quantity = max(*[len(item) for item in values[1:]])
    data = pandas.DataFrame(
        values[1:],
        columns=values[0]
        + [f"Undefined {item}" for item in range(columns_quantity - len(values[0]))],
    )

    data.rename(
        columns=dict(map(lambda item: (item, parse_str(item)), data.columns)),
        inplace=True,
    )
    data = data[PAYMENTS_COLUMNS.keys()]
    for column, parser in PAYMENTS_COLUMNS.items():
        data[column] = data[column].apply(parser)

    with open(Path(DATA_PATH / "payments.pkl"), "wb") as file_ref:
        pickle.dump(data, file_ref)


@log_execution_time("get_stats")
def get_stats():
    def processing_source(
        values: List[List[str]], columns_info: List[Tuple[str, str, Callable]]
    ) -> pandas.DataFrame:
        columns_count = max(list(map(lambda item: len(item), values[1:])))
        headers_count = len(values[0])
        if columns_count > headers_count:
            values[0] += list(
                map(
                    lambda item: f"Undefined {item}",
                    range(columns_count - headers_count),
                )
            )
        source = pandas.DataFrame(data=values[1:], columns=slugify_columns(values[0]))

        undefined_columns = list(
            set(map(lambda item: item[0], columns_info)) - set(source.columns)
        )
        assert len(undefined_columns) == 0, f"Undefined columns: {undefined_columns}"

        source = source[map(lambda item: item[0], columns_info)].rename(
            columns=dict(map(lambda item: item[0:2], columns_info))
        )
        parser_relation = dict(map(lambda item: item[1:3], columns_info))
        source = source.apply(lambda item: item.apply(parser_relation.get(item.name)))

        return source

    tilda = pickle_loader.leads
    tilda["amo_current"] = tilda["current_lead_amo"].apply(parse_lead_url)
    tilda["amo_main"] = tilda["main_lead_amo"].apply(parse_lead_url)
    tilda["target_short"] = tilda["traffic_channel"].apply(
        lambda item: urlparse(item, allow_fragments=False)
        ._replace(params=None, query=None, fragment=None)
        .geturl()
    )

    roistat_levels = pickle_loader.roistat_levels
    leads = pickle_loader.roistat_db
    channels = leads[["account"]].drop_duplicates(subset=["account"])
    channels["account_title"] = channels["account"].apply(
        lambda account_id: roistat_levels.loc[account_id]["title"]
    )
    channels["account"] = channels["account"].apply(
        lambda account_id: roistat_levels.loc[account_id]["name"]
    )
    channels.sort_values(by=["account"], inplace=True)
    channels.reset_index(drop=True, inplace=True)

    url_account = pickle_loader.roistat_leads[["url", "account", "qa1"]]
    url_account = url_account.merge(channels, how="left", on="account").drop_duplicates(
        subset=["url", "account", "account_title"], ignore_index=True
    )

    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        CREDENTIALS_FILE,
        [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    http_auth = credentials.authorize(Http())
    service = discovery.build("sheets", "v4", http=http_auth)
    spreadsheet_id = "1C4TnjTkSIsHs2svSgyFduBpRByA7M_i2sa6hrsX84EE"
    for sheet in (
        service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute().get("sheets")
    ):
        title = sheet.get("properties").get("title")
        values = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=title, majorDimension="ROWS")
            .execute()
        ).get("values")

        if title == "Все оплаты":
            source_payments: pandas.DataFrame = processing_source(
                values,
                [
                    ("menedzher", "manager", parse_str),
                    ("gr", "group", parse_str_with_empty),
                    ("ssylka_na_amocrm", "amo", parse_lead_url),
                    ("data_poslednej_zajavki_platnoj", "order_date", parse_date),
                    ("summa_vyruchki", "profit", parse_int),
                    ("data_oplaty", "profit_date", parse_date),
                    ("data_zoom", "zoom_date", parse_date),
                    ("tselevaja_ssylka", "target_link", parse_str),
                ],
            )
            source_payments["lead_id"] = source_payments["amo"].apply(parse_lead_id)
            for index, item in source_payments.iterrows():
                lead = detect_lead(item, tilda)
                channel_unique = ""
                channel = "Undefined"
                country = ""
                if lead is not None:
                    account_by_url = url_account[
                        url_account["url"] == lead["traffic_channel"]
                    ].reset_index(drop=True)
                    if not account_by_url.empty:
                        channel_unique = account_by_url.loc[0, "account"]
                        channel = account_by_url.loc[0, "account_title"]
                        country = account_by_url.loc[0, "qa1"]
                source_payments.loc[index, "channel"] = channel
                source_payments.loc[index, "channel_unique"] = channel_unique
                source_payments.loc[index, "country"] = country
            source_payments["channel"] = source_payments["channel"].apply(parse_str)
            source_payments["channel_id"] = source_payments["channel"].apply(parse_slug)
            source_payments["channel_unique"] = source_payments["channel_unique"].apply(
                parse_str
            )
            source_payments["country"] = (
                source_payments["country"].apply(parse_str).fillna("")
            )
            source_payments.insert(
                0, "manager_id", source_payments["manager"].apply(parse_slug)
            )
            source_payments = source_payments[
                ~(
                    source_payments["manager_id"].isna()
                    | (source_payments["manager_id"] == "")
                    | source_payments["lead_id"].isna()
                )
            ].reset_index(drop=True)
            source_payments["profit"].fillna(0, inplace=True)

        elif title == "SpecialOffers":
            source_so: pandas.DataFrame = processing_source(
                values,
                [
                    ("menedzher", "manager", parse_str),
                    ("gruppa", "group", parse_str_with_empty),
                    ("sdelka", "lead_id", parse_lead_id),
                    ("data_so", "so_date", parse_date),
                ],
            )
            source_so.insert(0, "manager_id", source_so["manager"].apply(parse_slug))
            source_so = source_so[
                ~(
                    source_so["manager_id"].isna()
                    | (source_so["manager_id"] == "")
                    | source_so["lead_id"].isna()
                )
            ].reset_index(drop=True)
            source_so.drop_duplicates(subset=["lead_id"], inplace=True)
            source_so.reset_index(drop=True, inplace=True)

    # --- Корректируем группу менеджера ----------------------------------------
    manager_group = pandas.concat(
        [
            source_payments[["manager_id", "manager", "group"]],
            source_so[["manager_id", "manager", "group"]],
        ],
        ignore_index=True,
    )
    groups_list = []
    for (manager_id, manager), manager_id_rows in manager_group.groupby(
        by=["manager_id", "manager"]
    ):
        groups_count = dict(
            map(
                lambda item: (item[0], len(item[1])),
                manager_id_rows.groupby(by=["group"]),
            )
        )
        groups_count = dict(filter(lambda item: item[0] != "", groups_count.items()))
        groups_count = sorted(
            groups_count.items(), key=lambda item: item[1], reverse=True
        )
        groups_list.append(
            [
                manager_id,
                manager,
                groups_count[0][0] if len(groups_count) else "",
            ]
        )
    groups = (
        pandas.DataFrame(groups_list, columns=["manager_id", "manager", "group"])
        .drop_duplicates(subset=["manager_id"])
        .sort_values(by=["manager"])
        .reset_index(drop=True)
    )
    source_payments.drop(columns=["group", "manager"], inplace=True)
    source_so.drop(columns=["group", "manager"], inplace=True)
    # --------------------------------------------------------------------------

    # --- Собираем количество лидов по каналам ---------------------------------
    channels_leads = pickle_loader.roistat_leads.rename(columns={"date": "datetime"})
    channels_leads.insert(
        0, "date", channels_leads["datetime"].apply(lambda item: item.date())
    )
    channels_leads = channels_leads.merge(channels, how="left", on="account")
    channels_leads["channel_id"] = channels_leads["account_title"].apply(parse_slug)

    channels_count_list: List[pandas.DataFrame] = []
    for (channel_id, date), items in channels_leads.groupby(by=["channel_id", "date"]):
        channels_count_list.append(
            {
                "channel_id": channel_id,
                "date": date,
                "count": len(items),
            }
        )
    channels_count = pandas.DataFrame(channels_count_list)

    channels_count_list_russia: List[pandas.DataFrame] = []
    for (channel_id, date), items in channels_leads[
        channels_leads["qa1"].str.contains("Россия", case=False)
    ].groupby(by=["channel_id", "date"]):
        channels_count_list_russia.append(
            {
                "channel_id": channel_id,
                "date": date,
                "count": len(items),
            }
        )
    channels_count_russia = pandas.DataFrame(channels_count_list_russia)
    # --------------------------------------------------------------------------

    # --- Собираем расходы -----------------------------------------------------
    roistat: pandas.DataFrame = leads.copy()
    roistat["date"] = roistat["date"].apply(lambda item: item.date())
    roistat = roistat[["date", "account", "expenses"]]
    roistat["account"] = roistat["account"].apply(
        lambda item: roistat_levels.loc[item]["name"]
    )
    roistat = roistat.merge(channels, how="left", on=["account"])
    roistat = roistat.rename(columns={"account_title": "channel", "expenses": "count"})
    roistat["count"] = roistat["count"].apply(parse_float)
    roistat = roistat[roistat["count"] > 0].reset_index(drop=True)
    roistat["channel_id"] = roistat["channel"].apply(parse_slug)
    roistat.drop(columns=["channel"], inplace=True)
    expenses_count_list = []
    for (channel_id, expenses_date), rows in roistat.groupby(by=["channel_id", "date"]):
        expenses_count_list.append([channel_id, expenses_date, rows["count"].sum()])
    expenses_count = pandas.DataFrame(
        expenses_count_list, columns=["channel_id", "date", "count"]
    )
    expenses_count["count"] = expenses_count["count"].apply(parse_int)
    expenses_count.sort_values(by=["channel_id", "date"], inplace=True)
    expenses = (
        source_payments[
            ~(
                source_payments["order_date"].isna()
                | source_payments["profit_date"].isna()
            )
        ][
            [
                "manager_id",
                "lead_id",
                "channel_id",
                "profit",
                "profit_date",
                "order_date",
                "country",
            ]
        ]
        .rename(columns={"order_date": "date"})
        .reset_index(drop=True)
    )
    # --------------------------------------------------------------------------

    # --- Собираем количество zoom ---------------------------------------------
    with open(DATA_PATH / "managers_zooms.pkl", "rb") as file_ref:
        source_zoom_count = pickle.load(file_ref)
    zoom_count_list = []
    for (manager_id, zoom_date), rows in source_zoom_count.groupby(
        by=["manager_id", "date"]
    ):
        zoom_count_list.append(
            {
                "manager_id": manager_id,
                "date": zoom_date,
                "count": len(rows),
            }
        )
    zoom_count = (
        pandas.DataFrame(zoom_count_list)
        .sort_values(by=["date"])
        .reset_index(drop=True)
    )
    # --------------------------------------------------------------------------

    # --- Собираем оплаты zoom -------------------------------------------------
    zoom = (
        source_payments[
            ~(
                source_payments["zoom_date"].isna()
                | source_payments["profit_date"].isna()
            )
        ][
            [
                "manager_id",
                "lead_id",
                "channel_id",
                "profit",
                "profit_date",
                "zoom_date",
            ]
        ]
        .rename(columns={"zoom_date": "date"})
        .reset_index(drop=True)
    )
    # --------------------------------------------------------------------------

    # --- Собираем количество so -----------------------------------------------
    so_count_list = []
    for (manager_id, so_date), rows in source_so.groupby(by=["manager_id", "so_date"]):
        so_count_list.append([manager_id, so_date, len(rows)])
    so_count = (
        pandas.DataFrame(so_count_list, columns=["manager_id", "date", "count"])
        .sort_values(by=["date"])
        .reset_index(drop=True)
    )
    # --------------------------------------------------------------------------

    # --- Собираем оплаты so ---------------------------------------------------
    so: pandas.DataFrame = (
        source_so.merge(source_payments, how="left", on=["manager_id", "lead_id"])[
            [
                "manager_id",
                "lead_id",
                "channel_id",
                "profit",
                "profit_date",
                "so_date",
            ]
        ]
        .rename(columns={"so_date": "date"})
        .reset_index(drop=True)
    )
    so = so[~so["profit_date"].isna()].reset_index(drop=True)
    so["profit"] = so["profit"].apply(parse_int)
    so["profit_date"] = so["profit_date"].apply(parse_date)
    so["date"] = so["date"].apply(parse_date)
    # --------------------------------------------------------------------------

    with open(Path(DATA_PATH / "groups.pkl"), "wb") as file_ref:
        pickle.dump(groups, file_ref)

    with open(Path(DATA_PATH / "channels.pkl"), "wb") as file_ref:
        pickle.dump(channels, file_ref)

    with open(Path(DATA_PATH / "channels_count.pkl"), "wb") as file_ref:
        pickle.dump(channels_count, file_ref)

    with open(Path(DATA_PATH / "channels_count_russia.pkl"), "wb") as file_ref:
        pickle.dump(channels_count_russia, file_ref)

    with open(Path(DATA_PATH / "expenses.pkl"), "wb") as file_ref:
        pickle.dump(expenses, file_ref)

    with open(Path(DATA_PATH / "expenses_count.pkl"), "wb") as file_ref:
        pickle.dump(expenses_count, file_ref)

    with open(Path(DATA_PATH / "zoom.pkl"), "wb") as file_ref:
        pickle.dump(zoom, file_ref)

    with open(Path(DATA_PATH / "zoom_count.pkl"), "wb") as file_ref:
        pickle.dump(zoom_count, file_ref)

    with open(Path(DATA_PATH / "so.pkl"), "wb") as file_ref:
        pickle.dump(so, file_ref)

    with open(Path(DATA_PATH / "so_count.pkl"), "wb") as file_ref:
        pickle.dump(so_count, file_ref)

    with open(Path(DATA_PATH / "source_payments.pkl"), "wb") as file_ref:
        pickle.dump(source_payments, file_ref)


@log_execution_time("update_so")
def update_so():
    def write_xlsx(data: List[List[Union[str, int]]]):
        workbook = Workbook(DATA_PATH / "so.xlsx")
        worksheet = workbook.add_worksheet("SpecialOffers")
        for index, row in enumerate(data):
            worksheet.write_row(index, 0, row)
        worksheet.autofilter(0, 0, len(data) - 1, len(data[0]) - 1)
        workbook.close()

    with open(Path(DATA_PATH / "so.pkl"), "rb") as file_ref:
        sources: pandas.DataFrame = pickle.load(file_ref)

    rel_fields = {
        "menedzher": parse_str,
        "gruppa": parse_str,
        "sdelka": parse_str,
        "lead_id": parse_lead_id,
        "fio_klienta": parse_str,
        "email": parse_str,
        "telefon": parse_str,
        "data_zoom": parse_date,
        "data_so": parse_date,
        "do_ili_posle_zoom": parse_str,
    }

    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        CREDENTIALS_FILE,
        [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    http_auth = credentials.authorize(Http())
    service = discovery.build("sheets", "v4", http=http_auth)
    spreadsheet_id = "1C4TnjTkSIsHs2svSgyFduBpRByA7M_i2sa6hrsX84EE"
    sheet_id = None
    data = None

    for sheet in (
        service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute().get("sheets")
    ):
        title = sheet.get("properties").get("title")
        if title == "SpecialOffers":
            sheet_id = sheet.get("properties").get("sheetId")
            values = (
                service.spreadsheets()
                .values()
                .get(spreadsheetId=spreadsheet_id, range=title, majorDimension="ROWS")
                .execute()
            )
            items = values.get("values", [])
            if len(items):
                items[0] = list(
                    map(lambda item: slugify(item, "ru").replace("-", "_"), items[0])
                )
                data = pandas.DataFrame(columns=items[0], data=items[1:])
                data.insert(3, "lead_id", data["sdelka"])
                for column, parse_fn in rel_fields.items():
                    if column not in data.columns:
                        data[column] = ""
                    data[column] = data[column].apply(parse_fn)
            break

    if data is not None:
        data = data[rel_fields.keys()]
        for index, row in data.iterrows():
            if not pandas.isna(row["lead_id"]):
                row_sources = sources[sources["lead_id"] == row["lead_id"]].sort_values(
                    by=["profit_date"]
                )
                num = 1
                for _, payment in row_sources.iterrows():
                    data.loc[index, [f"data_oplaty_{num}", f"summa_oplaty_{num}"]] = [
                        payment["profit_date"],
                        payment["profit"],
                    ]
                    num += 1

        if sheet_id is not None:
            data.drop(columns=["lead_id"], inplace=True)
            data.fillna("", inplace=True)
            for column in data.columns:
                if column in ["data_zoom", "data_so"] or str(column).startswith(
                    "data_oplaty_"
                ):
                    data[column] = data[column].apply(
                        lambda item: str(item.strftime("%d.%m.%Y")) if item else ""
                    )
                elif str(column).startswith("summa_oplaty_"):
                    data[column] = data[column].apply(
                        lambda item: "" if str(item) == "" else int(item)
                    )
                else:
                    data[column] = data[column].apply(lambda item: f"'{item}")
            data.rename(
                columns=dict(
                    zip(data.columns, list(map(rename_so_columns, data.columns)))
                ),
                inplace=True,
            )
            values = [list(data.columns)] + data.values.tolist()
            write_xlsx(values)
            requests_data = [
                {
                    "deleteSheet": {
                        "sheetId": sheet_id,
                    }
                },
                {
                    "addSheet": {
                        "properties": {
                            "title": "SpecialOffers",
                        }
                    }
                },
            ]
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id, body={"requests": requests_data}
            ).execute()
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"SpecialOffers!A1:ZZ{len(data)+1}",
                valueInputOption="USER_ENTERED",
                body={"values": values},
            ).execute()


@log_execution_time("get_managers_zooms")
def get_managers_zooms():
    service = get_google_service()

    spreadsheet = get_spreadsheets_month(service)

    data = None
    managers = pandas.DataFrame(columns=["manager", "group", "date"])
    for _, row in spreadsheet.iterrows():
        managers_month, data_month = read_zoom_month(
            spreadsheet_id=row["hash"],
            month=datetime.strptime(row["month"], "%Y%m").date(),
        )
        if managers_month is not None:
            managers = pandas.concat([managers, managers_month]).reset_index(drop=True)
        if data_month is None:
            continue
        if data is None:
            data = pandas.DataFrame()
        data = pandas.concat([data, data_month])
    unique_managers = []
    for manager_name, manager in managers[["manager", "group"]].groupby(by=["manager"]):
        groups_count = manager.groupby(by=["group"]).count().to_dict().get("manager")
        groups_keys, groups_values = list(groups_count.keys()), list(
            groups_count.values()
        )
        if not groups_values:
            group = ""
        else:
            group = groups_keys[
                groups_values.index(sorted(groups_values, reverse=True)[0])
            ]
        unique_managers.append({"manager": manager_name, "group": group})
    unique_managers = pandas.DataFrame(unique_managers)
    unique_managers["manager_id"] = unique_managers["manager"].apply(parse_slug)
    managers = managers[["manager", "date"]].merge(
        unique_managers, how="left", on="manager"
    )

    data["manager_id"] = data["manager"].apply(parse_slug)

    data = data.merge(
        unique_managers[["manager_id", "group"]], how="left", on="manager_id"
    )
    data.rename(columns={"group": "group_id"}, inplace=True)
    data["group_id"].fillna("", inplace=True)
    data["group"] = data["group_id"].apply(lambda item: f'Группа "{item}"')

    payments = read_payments(service, "1C4TnjTkSIsHs2svSgyFduBpRByA7M_i2sa6hrsX84EE")
    data = data.merge(payments, how="left", on=["manager_id", "lead", "date"])
    data["profit"] = data["profit"].fillna(0).apply(parse_int)

    with open(Path(DATA_PATH / "managers_zooms.pkl"), "wb") as file_ref:
        pickle.dump(data, file_ref)
    with open(Path(DATA_PATH / "managers_groups.pkl"), "wb") as file_ref:
        pickle.dump(managers, file_ref)


@log_execution_time("get_managers_sales")
def get_managers_sales():
    response = requests.get(
        "https://docs.google.com/spreadsheets/d/1ZOsLTSWLGIFwZfWusioUcyngGXE_sGSSbqs5-ELsg_w/export?format=xlsx&id=1ZOsLTSWLGIFwZfWusioUcyngGXE_sGSSbqs5-ELsg_w"
    )
    groups: pandas.DataFrame = pandas.read_excel(
        BytesIO(response.content), sheet_name="Продукты УИИ", dtype=str
    )
    groups.rename(
        columns=dict(zip(list(groups.columns), slugify_columns(list(groups.columns)))),
        inplace=True,
    )
    groups["nazvanie_produkta"] = groups["nazvanie_produkta"].apply(parse_str)
    groups["gruppa_produktov"] = groups["gruppa_produktov"].apply(parse_str)
    groups.rename(
        columns={"nazvanie_produkta": "course", "gruppa_produktov": "group"},
        inplace=True,
    )
    groups = groups[groups["group"] != "не считаем"].reset_index(drop=True)
    data: pandas.DataFrame = pandas.read_pickle(Path(DATA_PATH / "payments.pkl"))
    data.rename(
        columns=dict(zip(list(data.columns), slugify_columns(list(data.columns)))),
        inplace=True,
    )
    data["lead_id"] = data["ssylka_na_amocrm"].apply(parse_lead_id)
    columns_rel = {
        "data_oplaty": "payment_date",
        "data_poslednej_zajavki_platnoj": "order_date",
        "menedzher": "manager",
        "lead_id": "lead",
        "kurs": "course",
        "summa_vyruchki": "profit",
        "mesjats_doplata": "surcharge",
    }
    data = data[columns_rel.keys()].rename(columns=columns_rel)
    data = data[~data["lead"].isnull()]

    data["manager"] = data["manager"].apply(parse_str).fillna("undefined")
    data["course"] = data["course"].apply(parse_str).fillna("undefined")
    data["profit"] = data["profit"].fillna(0).apply(parse_float)
    data["surcharge"] = (
        data["surcharge"].apply(lambda item: item == "доплата").fillna(False)
    )

    for _, payments in data.groupby(by=["lead", "course"]):
        surcharge = payments[payments["surcharge"]]
        not_surcharge = payments[~payments["surcharge"]]
        if not surcharge.empty:
            if not not_surcharge.empty:
                data.loc[surcharge.index, "payment_date"] = not_surcharge.reset_index(
                    drop=True
                ).iloc[0]["payment_date"]
        if not_surcharge.empty:
            data.drop(index=payments.index, inplace=True)

    data = data.merge(groups, how="left", on="course")
    data["group"] = data["group"].fillna("Undefined")
    data.drop(columns=["surcharge", "lead"], inplace=True)
    data.reset_index(drop=True, inplace=True)

    with open(Path(DATA_PATH / "managers_sales.pkl"), "wb") as file_ref:
        pickle.dump(data, file_ref)


@log_execution_time("get_funnel_channel")
def get_funnel_channel():
    rels = {
        "Интенсив 3 дня": [
            "https://neural-university.ru/lp",
            "https://neural-university.ru/lp_pro",
            "https://neural-university.ru/lp_direct",
            "https://terrauniversity.site",
            "https://neural-university.ru/lp_marketing",
            "https://neural-university.ru/lp_trading",
            "https://neural-university.ru/lp_medicine",
            "https://neural-university.ru/lp_sysadministrator",
            "https://neural-university.ru/lp_1c_programmer",
            "https://neural-university.ru/lp_project_product_manager",
            "https://neural-university.ru/lp_web_programmer",
            "https://neural-university.ru/lp_e-commerce",
            "https://neural-university.ru/lp_vk",
            "https://neural-university.ru/lp1",
            "https://neural-university.ru/lp2",
            "https://neural-university.ru/lp3",
            "https://neural-university.ru/lp4",
            "https://neural-university.ru/lp_python",
            "https://neural-university.ru/lp_java",
            "https://terrauniversity.website",
            "https://terra-university.website",
            "https://neural-university.ru/lp_protech",
            "https://neural-university.ru/lp_mka",
            "https://neural-university.ru/lp_pro",
            "https://neural-university.ru/lp_mask",
            "https://neural-university.ru/lp_engineer",
        ],
        "Интенсив 2 дня": [
            "https://neural-university.ru/lp_2day",
            "https://neural-university.ru/web_16052023",
            "https://ai-university.ru/lp_2day",
            "https://ai-university.ru/",
        ],
        "ChatGPT": [
            "https://neural-university.ru/lp_chatgpt_web",
            "https://neural-university.ru/webinar_chatgpt",
            "https://neural-university.ru/lp_chatgpt_course",
            "https://neural-university.ru/chatgpt_freecourse",
            "https://neural-university.ru/lp_chatgpt_neurostaff_web",
        ],
        "Курс 7 уроков": ["https://neural-university.ru/free_course"],
    }

    def parse_url_path(value: str):
        url = urlparse(str(value))
        if url.scheme and url.netloc and url.path:
            return f"{url.scheme}://{url.netloc}{url.path}"
        return pandas.NA

    def parse_funnel(value: str):
        items = list(dict(filter(lambda item: value in item[1], rels.items())).keys())
        items.append(pandas.NA)
        return items[0]

    with open(DATA_PATH / "source_payments.pkl", "rb") as file_ref:
        data: pandas.DataFrame = pandas.read_pickle(file_ref)
    data = data[
        [
            "order_date",
            "channel_unique",
            "channel",
            "profit_date",
            "profit",
            "target_link",
        ]
    ].rename(
        columns={
            "order_date": "date",
            "channel_unique": "account",
            "channel": "account_title",
            "target_link": "url",
        }
    )
    data["date"] = data["date"].apply(parse_date)
    data = data[~data["date"].isna()]
    data["profit_date"] = data["profit_date"].apply(parse_date)
    data = data[~data["profit_date"].isna()]
    data["url"] = data["url"].apply(parse_url_path)
    data = data[~data["url"].isna()]
    data["funnel"] = data["url"].apply(parse_funnel)
    data = data[~data["funnel"].isna()]
    data = data[["date", "funnel", "account", "account_title", "profit_date", "profit"]]
    data["account"].fillna("undefined", inplace=True)
    data["account_title"].fillna("Undefined", inplace=True)

    roistat_levels = pickle_loader.roistat_levels
    channels = pickle_loader.roistat_db[["account"]].drop_duplicates(subset=["account"])
    channels["account_title"] = channels["account"].apply(
        lambda account_id: roistat_levels.loc[account_id]["title"]
    )
    channels["account"] = channels["account"].apply(
        lambda account_id: roistat_levels.loc[account_id]["name"]
    )
    channels.reset_index(drop=True, inplace=True)
    expenses = pickle_loader.roistat_leads[["account", "url", "expenses", "date"]]
    expenses = expenses[expenses["date"].apply(lambda item: isinstance(item, datetime))]
    expenses["date"] = expenses["date"].apply(lambda item: item.date())
    expenses["url"] = expenses["url"].apply(parse_url_path)
    expenses = expenses[~expenses["url"].isna()]
    expenses["funnel"] = expenses["url"].apply(parse_funnel)
    expenses = expenses[~expenses["funnel"].isna()]
    expenses = expenses[["date", "funnel", "account", "expenses"]]
    expenses = expenses.merge(channels, how="left", on="account")

    rows_expenses = []
    for (date, funnel, account, account_title), group in expenses.groupby(
        by=["date", "funnel", "account", "account_title"]
    ):
        rows_expenses.append(
            {
                "date": date,
                "funnel": funnel,
                "channel": account,
                "channel_title": account_title,
                "expenses": group["expenses"].sum(),
            }
        )

    rows_profit = []
    for (date, funnel, account, account_title, profit_date), group in data.groupby(
        by=["date", "funnel", "account", "account_title", "profit_date"]
    ):
        rows_profit.append(
            {
                "date": date,
                "funnel": funnel,
                "channel": account,
                "channel_title": account_title,
                "profit_date": profit_date,
                "profit": group["profit"].sum(),
            }
        )

    with open(Path(DATA_PATH / "funnel_channel_expenses.pkl"), "wb") as file_ref:
        pickle.dump(pandas.DataFrame(data=rows_expenses), file_ref)

    with open(Path(DATA_PATH / "funnel_channel_profit.pkl"), "wb") as file_ref:
        pickle.dump(pandas.DataFrame(data=rows_profit), file_ref)


@log_execution_time("get_intensives_emails")
def get_intensives_emails():
    data_columns = ["course", "date", "email"]

    def sources_read(url: Optional[str] = None) -> Optional[pandas.ExcelFile]:
        if url is None:
            return
        print(f"   | Read source from url: {url}")
        response = requests.get(url)
        try:
            return pandas.ExcelFile(BytesIO(response.content))
        except ValueError:
            print(f"     | Incorrect type of source file")

    def get_download_url(url: str) -> Optional[str]:
        print(f"   | Download file from url: {url}")
        try:
            url = urlparse(url)
            tails = list(filter(None, url.path.split("/")))
            return f"https://docs.google.com/spreadsheets/d/{tails[2]}/export?format=xlsx&id={tails[2]}"
        except TypeError as error:
            print(f"     | URL is undefined")

    def read_excel(
        course_name: str, file: Optional[pandas.ExcelFile]
    ) -> Optional[pandas.DataFrame]:
        if file is None:
            return
        print(f"   | Read excel document: {file.sheet_names}")
        dataframe = pandas.DataFrame(columns=data_columns)
        for sheet_name in file.sheet_names:
            try:
                date = datetime.strptime(sheet_name, "%d.%m.%Y").date()
            except ValueError:
                continue
            data: pandas.DataFrame = file.parse(sheet_name)
            if data.empty:
                data = pandas.DataFrame(columns=data_columns)
            data.rename(
                columns=dict(zip(data.columns, slugify_columns(list(data.columns)))),
                inplace=True,
            )
            data.rename(columns={"e_mail": "email"}, inplace=True)
            data.drop_duplicates(inplace=True, ignore_index=True)
            data["email"] = data["email"].apply(parse_str)
            data["date"] = date
            data["course"] = course_name
            dataframe = pandas.concat([dataframe, data], ignore_index=True)
        return dataframe[data_columns]

    print("-> Read summary")
    summary_file = sources_read(
        get_download_url(
            "https://docs.google.com/spreadsheets/d/1KdI82fdMge4PQ3FqLfQkYdUj28ogMLhh/edit"
        )
    )
    print()
    if summary_file is None:
        return

    summary: pandas.DataFrame = summary_file.parse()
    summary.rename(
        columns=dict(
            zip(list(summary.columns), slugify_columns(list(summary.columns)))
        ),
        inplace=True,
    )
    summary.rename(
        columns={
            "fajl_s_registratsijami": "intensives_registrations",
            "fajl_s_predzakazami": "intensives_preorders",
        },
        inplace=True,
    )
    for column in summary.columns:
        summary[column] = summary[column].apply(parse_str)
    sources = {
        "intensives_registrations": pandas.DataFrame(columns=data_columns),
        "intensives_preorders": pandas.DataFrame(columns=data_columns),
    }
    for _, summary_row in summary.iterrows():
        print("-> Read sources")
        registrations_dataframe = read_excel(
            summary_row["intensiv"],
            sources_read(get_download_url(summary_row.intensives_registrations)),
        )
        if registrations_dataframe is not None:
            sources["intensives_registrations"] = pandas.concat(
                [sources["intensives_registrations"], registrations_dataframe],
                ignore_index=True,
            )
        print("   |")

        preorders_dataframe = read_excel(
            summary_row["intensiv"],
            sources_read(get_download_url(summary_row.intensives_preorders)),
        )
        if preorders_dataframe is not None:
            sources["intensives_preorders"] = pandas.concat(
                [sources["intensives_preorders"], preorders_dataframe],
                ignore_index=True,
            )
        print()

    for source_name, source in sources.items():
        with open(Path(DATA_PATH / f"{source_name}.pkl"), "wb") as file_ref:
            pickle.dump(source, file_ref)


@log_execution_time("get_intensives_so")
def get_intensives_so():
    data_columns = ["date", "email"]
    response = requests.get(
        "https://docs.google.com/spreadsheets/d/11sZcWN74isA-4zC73HDteuwR_ZDXMiNKkR2fsyLdi74/export?format=xlsx&id=11sZcWN74isA-4zC73HDteuwR_ZDXMiNKkR2fsyLdi74"
    )
    data_file = pandas.ExcelFile(BytesIO(response.content))
    dataframe = pandas.DataFrame(columns=data_columns)
    for sheet_name in data_file.sheet_names:
        try:
            date = datetime.strptime(sheet_name, "%d.%m.%Y").date()
        except ValueError:
            continue
        data: pandas.DataFrame = data_file.parse(sheet_name)
        if data.empty:
            data = pandas.DataFrame(columns=data_columns)
        data.rename(
            columns=dict(zip(data.columns, slugify_columns(list(data.columns)))),
            inplace=True,
        )
        data.rename(columns={"e_mail": "email"}, inplace=True)
        if "email" not in list(data.columns):
            continue
        data.drop_duplicates(inplace=True, ignore_index=True)
        data["email"] = data["email"].apply(parse_str)
        data["date"] = date
        dataframe = pandas.concat([dataframe, data], ignore_index=True)
    data = dataframe[data_columns]

    with open(Path(DATA_PATH / f"intensives_so.pkl"), "wb") as file_ref:
        pickle.dump(data, file_ref)


@log_execution_time("get_source_so")
def get_source_so():
    response = requests.get(
        "https://docs.google.com/spreadsheets/d/1C4TnjTkSIsHs2svSgyFduBpRByA7M_i2sa6hrsX84EE/export?format=xlsx&id=1C4TnjTkSIsHs2svSgyFduBpRByA7M_i2sa6hrsX84EE"
    )
    data_file = pandas.ExcelFile(BytesIO(response.content))
    data: pandas.DataFrame = data_file.parse("SpecialOffers")
    data.rename(
        columns=dict(zip(data.columns, slugify_columns(list(data.columns)))),
        inplace=True,
    )
    data_list = []
    for index, row in data.iterrows():
        row_data = {
            "manager": row["menedzher"],
            "group": row["gruppa"],
            "lead": row["sdelka"],
            "client": row["fio_klienta"],
            "email": row["email"],
            "phone": row["telefon"],
            "date": row["data_so"],
        }
        is_payment = False
        for column in data.columns:
            if not str(column).startswith("data_oplaty_"):
                continue
            if not pandas.isna(row[column]):
                is_payment = True
                data_list.append(
                    {
                        **row_data,
                        "payment_date": row[column],
                        "payment": row[f"summa_oplaty_{column[12:]}"],
                    }
                )
        if not is_payment:
            data_list.append(row_data)

    data = pandas.DataFrame(data_list)
    data["manager"] = data["manager"].apply(parse_str)
    data["group"] = data["group"].apply(parse_str)
    data["lead"] = data["lead"].apply(parse_lead_url)
    data["lead_id"] = data["lead"].apply(parse_lead_id)
    data["client"] = data["client"].apply(parse_str)
    data["email"] = data["email"].apply(parse_str)
    data["phone"] = data["phone"].apply(parse_str)
    data["date"] = data["date"].apply(parse_date)
    data["payment_date"] = data["payment_date"].apply(parse_date)
    data["payment"] = data["payment"].apply(parse_float)

    with open(Path(DATA_PATH / f"source_so.pkl"), "wb") as file_ref:
        pickle.dump(data, file_ref)


dag = DAG(
    "week_stats",
    description="Collect week statistics",
    schedule_interval="15 6-19 * * *",
    start_date=datetime(2017, 3, 20),
    catchup=False,
)


get_payments_operator = PythonOperator(
    task_id="get_payments",
    python_callable=get_payments,
    dag=dag,
)
get_source_so_operator = PythonOperator(
    task_id="get_source_so",
    python_callable=get_source_so,
    dag=dag,
)
get_stats_operator = PythonOperator(
    task_id="get_stats",
    python_callable=get_stats,
    dag=dag,
)
update_so_operator = PythonOperator(
    task_id="update_so",
    python_callable=update_so,
    dag=dag,
)
get_managers_zooms_operator = PythonOperator(
    task_id="get_managers_zooms",
    python_callable=get_managers_zooms,
    dag=dag,
)
get_managers_sales_operator = PythonOperator(
    task_id="get_managers_sales",
    python_callable=get_managers_sales,
    dag=dag,
)
get_intensives_emails_operator = PythonOperator(
    task_id="get_intensives_emails",
    python_callable=get_intensives_emails,
    dag=dag,
)
get_intensives_so_operator = PythonOperator(
    task_id="get_intensives_so",
    python_callable=get_intensives_so,
    dag=dag,
)
get_funnel_channel_operator = PythonOperator(
    task_id="get_funnel_channel",
    python_callable=get_funnel_channel,
    dag=dag,
)

get_payments_operator >> get_managers_sales_operator
get_managers_zooms_operator >> get_stats_operator
get_stats_operator >> update_so_operator
get_stats_operator >> get_funnel_channel_operator
